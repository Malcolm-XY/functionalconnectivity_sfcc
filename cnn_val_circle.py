# -*- coding: utf-8 -*-
"""
Created on Fri Jan  3 14:34:33 2025

@author: 18307
"""

import os
import pandas as pd

import utils
import cnn_validation
import covmap_construct

def cnn_validation_circle(model, method, feature, subject_range, experiment_range):
    labels = utils.get_label()
    
    results_entry = []
    for sub in subject_range:
        for ex in experiment_range:
            identifier = f'sub{sub}ex{ex}'
            print(f'Processing {identifier}...')
            
            # Get cm data
            cmdata = utils.load_cmdata2d(feature, 'joint', identifier)
            
            if method == 'sfcc':
                # Draw sfcc
                fcdata = covmap_construct.generate_sfcc(cmdata, imshow=True)
            elif method == 'cm':
                fcdata = cmdata
            
            # Validation
            result = cnn_validation.cnn_validation(model, fcdata, labels)
            
            # Add identifier to the result
            result['Identifier'] = f'sub{sub}ex{ex}'
            results_entry.append(result)

    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation compelete\n')
    
    return results_entry

def cnn_cross_validation_circle(model, method, feature, subject_range, experiment_range):
    labels = utils.get_label()
    
    results_entry = []
    for sub in subject_range:
        for ex in experiment_range:
            identifier = f'sub{sub}ex{ex}'
            print(f'Processing {identifier}...')

            # Get cm data
            cmdata = utils.load_cmdata2d(feature, 'joint', identifier, imshow=True)
            
            if method == 'sfcc':
                # Draw sfcc
                fcdata = covmap_construct.generate_sfcc(cmdata, imshow=True)
            elif method == 'cm':
                fcdata = cmdata
            
            # Validation
            result = cnn_validation.cnn_cross_validation(model, fcdata, labels)
            
            # Add identifier to the result
            result['Identifier'] = f'sub{sub}ex{ex}'
            results_entry.append(result)

    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation compelete\n')
    
    return results_entry

from openpyxl import load_workbook

def save_results_to_xlsx_append(results, output_dir, filename, sheet_name='K-Fold Results'):
    """
    Appends results to an existing Excel file or creates a new file if it doesn't exist.

    Args:
        results (list or pd.DataFrame): The results data to save.
        output_dir (str): The directory where the Excel file will be saved.
        filename (str): The name of the Excel file.
        sheet_name (str): The sheet name in the Excel file. Default is 'K-Fold Results'.

    Returns:
        str: The path of the saved Excel file.
    """
    # Convert results to DataFrame if necessary
    if not isinstance(results, pd.DataFrame):
        results_df = pd.DataFrame(results)
    else:
        results_df = results

    # Rearrange columns if "Identifier" is present
    if 'Identifier' in results_df.columns:
        columns_order = ['Identifier'] + [col for col in results_df.columns if col != 'Identifier']
        results_df = results_df[columns_order]

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the full output path
    output_path = os.path.join(output_dir, filename)

    # Append to existing Excel file or create a new one
    if os.path.exists(output_path):
        print(f"Appending data to existing file: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the existing workbook
            existing_workbook = load_workbook(output_path)

            # Check if the sheet exists
            if sheet_name in existing_workbook.sheetnames:
                # Load existing sheet and append
                start_row = existing_workbook[sheet_name].max_row
                results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
            else:
                # Write new sheet if not exists
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        print(f"Creating new file: {output_path}")
        results_df.to_excel(output_path, index=False, sheet_name=sheet_name)

    print(f"Results successfully saved to: {output_path}")
    return output_path

import time
import threading

def shutdown_with_countdown(countdown_seconds=30):
    """
    Initiates a shutdown countdown, allowing the user to cancel shutdown within the given time.

    Args:
        countdown_seconds (int): The number of seconds to wait before shutting down.
    """
    def cancel_shutdown():
        nonlocal shutdown_flag
        user_input = input("\nPress 'c' and Enter to cancel shutdown: ").strip().lower()
        if user_input == 'c':
            shutdown_flag = False
            print("Shutdown cancelled.")

    # Flag to determine whether to proceed with shutdown
    shutdown_flag = True

    # Start a thread to listen for user input
    input_thread = threading.Thread(target=cancel_shutdown, daemon=True)
    input_thread.start()

    # Countdown timer
    print(f"Shutdown scheduled in {countdown_seconds} seconds. Press 'c' to cancel.")
    for i in range(countdown_seconds, 0, -1):
        print(f"Time remaining: {i} seconds", end="\r")
        time.sleep(1)

    # Check the flag after countdown
    if shutdown_flag:
        print("\nShutdown proceeding...")
        os.system("shutdown /s /t 1")  # Execute shutdown command
    else:
        print("\nShutdown aborted.")

def end_program_actions(play_sound=True, shutdown=False, countdown_seconds=30):
    """
    Performs actions at the end of the program, such as playing a sound or shutting down the system.

    Args:
        play_sound (bool): If True, plays a notification sound.
        shutdown (bool): If True, initiates shutdown with a countdown.
        countdown_seconds (int): Countdown time for shutdown confirmation.
    """
    if play_sound:
        try:
            import winsound
            print("Playing notification sound...")
            winsound.Beep(1000, 500)  # Frequency: 1000Hz, Duration: 500ms
        except ImportError:
            print("winsound module not available. Skipping sound playback.")

    if shutdown:
        shutdown_with_countdown(countdown_seconds)

# %% Usage
from Models import models, models_multiscale

# training settings
model = models.CNN2DModel()
fcnetwork, feature, subject_range, experiment_range = 'sfcc', 'PCC', range(1, 2), range(1, 2)

# trainning and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, subject_range, experiment_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'Results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# End program actions
end_program_actions(play_sound=True, shutdown=True, countdown_seconds=60)