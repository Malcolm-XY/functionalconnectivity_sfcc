# -*- coding: utf-8 -*-
"""
Created on Fri Jan  3 14:34:33 2025

@author: 18307
"""

import os
import numpy as np
import pandas as pd

import torch

import cnn_validation
import utils_feature_loading
import utils_visualization

import covmap_construct
import rearrangedmap_construct

# %% Validation Circle
def cnn_validation_circle(model, fcnetwork, feature, subject_range, experiment_range):
    # labels and targets
    labels = utils_feature_loading.read_labels('seed')
    labels_tensor = torch.tensor(labels)
    unique_classes, targets = torch.unique(labels_tensor, sorted=True, return_inverse=True)
    
    results_entry = []
    for sub in subject_range:
        for ex in experiment_range:
            identifier = f'sub{sub}ex{ex}'
            print(f'Processing {identifier}...')
            
            # get connectivity matrices
            fcs = utils_feature_loading.read_fcs('seed', identifier, feature, band='joint')

            # feature engineering; functional connectivity networks
            if fcnetwork == 'sfcc':
                # Draw sfcc
                fcs = covmap_construct.generate_sfcc(fcs, 'SEED', imshow=False)
                fcs = np.stack(list(fcs.values()), axis=1)
                utils_visualization.draw_projection(np.mean(fcs, axis=0))

            elif fcnetwork == 'cm':
                fcs = fcs
                fcs = rearrangedmap_construct.global_padding(fcs)
                fcs = np.stack(list(fcs.values()), axis=1)
                utils_visualization.draw_projection(np.mean(fcs, axis=0))
            elif fcnetwork == 'mx':
                fcs = rearrangedmap_construct.generate_rearranged_fcs(fcs, 'MX', imshow = True)
            elif fcnetwork == 'vc':
                fcs = rearrangedmap_construct.generate_rearranged_fcs(fcs, 'VC', imshow = True)
            
            # Training and Validation
            result = cnn_validation.cnn_validation(model, fcs, targets)
            
            # Add identifier to the result
            result['Identifier'] = f'sub{sub}ex{ex}'
            results_entry.append(result)

    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation complete\n')
    
    return results_entry

# %% Cross Validation Circle
def cnn_cross_validation_circle(model, method, feature, subject_range, experiment_range):
    # labels and targets
    labels = np.array(utils_feature_loading.read_labels('seed'))
    labels = np.reshape(labels, -1)
    labels_tensor = torch.tensor(labels)
    unique_classes, targets = torch.unique(labels_tensor, sorted=True, return_inverse=True)
    
    results_entry = []
    for sub in subject_range:
        for ex in experiment_range:
            identifier = f'sub{sub}ex{ex}'
            print(f'Processing {identifier}...')

            # get connectivity matrices
            # fcs = utils_feature_loading.read_fcs('seed', identifier, feature, band='joint')
            fcs = utils_feature_loading.read_fcs_mat('seed', identifier, feature, 'joint')
            
            # feature engineering; functional connectivity networks
            if method == 'sfcc':
                # Draw sfcc
                fcs = covmap_construct.generate_sfcc(fcs, "SEED", imshow=False)
                fcs = np.stack([fcs['alpha'], fcs['beta'], fcs['gamma']], axis=1)

                utils_visualization.draw_projection(np.mean(fcs, axis=0))
            elif method == 'cm':
                fcs = fcs
                fcs = np.stack([fcs['alpha'], fcs['beta'], fcs['gamma']], axis=1)
                fcs = rearrangedmap_construct.global_padding(fcs)
                
                utils_visualization.draw_projection(np.mean(fcs, axis=0))
            elif method == 'mx':
                fcs = rearrangedmap_construct.generate_rearranged_fcs(fcs, 'MX', imshow = True)
            elif method == 'vc':
                fcs = rearrangedmap_construct.generate_rearranged_fcs(fcs, 'VC', imshow = True)
            
            # Validation
            result = cnn_validation.cnn_cross_validation(model, fcs, targets)
            
            # Add identifier to the result
            result['Identifier'] = f'sub{sub}ex{ex}'
            results_entry.append(result)

    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation complete\n')
    
    return results_entry

# %% Save Result Actions
from openpyxl import load_workbook
def save_results_to_xlsx_append(results, output_dir, filename, sheet_name='K-Fold Results'):
    """
    Appends results to an existing Excel file or creates a new file if it doesn't exist.

    Args:
        results (list or pd.DataFrame): The results data to save.
        output_dir (str): The directory where the Excel file will be saved.
        filename (str): The name of the Excel file.
        sheet_name (str): The sheet name in the Excel file. Default is 'K-Fold Results'.

    Returns:
        str: The path of the saved Excel file.
    """
    # Convert results to DataFrame if necessary
    if not isinstance(results, pd.DataFrame):
        results_df = pd.DataFrame(results)
    else:
        results_df = results

    # Rearrange columns if "Identifier" is present
    if 'Identifier' in results_df.columns:
        columns_order = ['Identifier'] + [col for col in results_df.columns if col != 'Identifier']
        results_df = results_df[columns_order]

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the full output path
    output_path = os.path.join(output_dir, filename)

    # Append to existing Excel file or create a new one
    if os.path.exists(output_path):
        print(f"Appending data to existing file: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the existing workbook
            existing_workbook = load_workbook(output_path)

            # Check if the sheet exists
            if sheet_name in existing_workbook.sheetnames:
                # Load existing sheet and append
                start_row = existing_workbook[sheet_name].max_row
                results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
            else:
                # Write new sheet if not exists
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        print(f"Creating new file: {output_path}")
        results_df.to_excel(output_path, index=False, sheet_name=sheet_name)

    print(f"Results successfully saved to: {output_path}")
    return output_path

# %% Usage; Training settings
from models import models #, models_multiscale

# model = models.CNN_2layers_adaptive_maxpool_3()
model = models.MSCNN_3_2layers_cv_235_adaptive_maxpool_3()

# %% validation 1; sfcc
fcnetwork, feature, subject_range, experiment_range = 'sfcc', 'PLV', range(10, 16), range(1, 4)

# training and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, subject_range, experiment_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% validation 2; cm
fcnetwork, feature, subject_range, experiment_range = 'cm', 'PLV', range(10, 16), range(1, 4)

# training and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, subject_range, experiment_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% validation 3; vc
fcnetwork, feature, subject_range, experiment_range = 'vc', 'PLV', range(1, 16), range(1, 4)

# training and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, subject_range, experiment_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% validation 4; mx
fcnetwork, feature, subject_range, experiment_range = 'mx', 'PLV', range(1, 16), range(1, 4)

# training and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, subject_range, experiment_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% End program actions
utils_visualization.end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120)