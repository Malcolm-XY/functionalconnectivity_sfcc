# -*- coding: utf-8 -*-
"""
Created on Sun Feb  2 15:28:38 2025

@author: 18307
"""

import os
import numpy as np
import pandas as pd
import torch

import cnn_validation

import utils_common
import featureengineering_dreamer
import covmap_construct
import rearrangedmap_construct

# %% Cross Validation Circle
def cnn_cross_validation_circle(model, fcnetwork, feature, emotion_dimension='arousal', subject_range=range(1,2)):
    # labels and targets
    labels = utils_common.read_labels_dreamer()
    labels = labels[emotion_dimension]
    labels_tensor = torch.tensor(labels)
    unique_classes, targets = torch.unique(labels_tensor, sorted=True, return_inverse=True)
    
    results_entry = []
    for subject in subject_range:
        identifier = f'sub{subject}'
        print(f'Processing {identifier}...')
        
        # get connectivity matrices       
        cms = utils_common.load_cms(dataset='DREAMER', experiment=identifier, feature=feature, band='joint', imshow=True)
        
        # feature engineering; functional connectivity networks
        if fcnetwork == 'sfcc':
            # Draw sfcc
            fcs = covmap_construct.generate_sfcc(cms, "DREAMER", imshow=False)
            fcs = featureengineering_dreamer.interpolate_matrices(fcs)
            utils_common.draw_projection(np.mean(fcs, axis=0))
        elif fcnetwork == 'cm':
            fcs = cms
            fcs = rearrangedmap_construct.global_padding(fcs)
            utils_common.draw_projection(np.mean(fcs, axis=0))
        elif fcnetwork == 'mx':
            fcs = rearrangedmap_construct.generate_rearrangedcm(cms, 'MX', order="DREAMER", padding=True, imshow = True)
        elif fcnetwork == 'vc':
            fcs = rearrangedmap_construct.generate_rearrangedcm(cms, 'VC', order="DREAMER", padding=True, imshow = True)
        
        # Validation
        result = cnn_validation.cnn_cross_validation(model, fcs, targets)
        
        # Add identifier to the result
        result['Identifier'] = f'sub{subject}'
        results_entry.append(result)

    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation compelete\n')
    
    return results_entry

# %% Save Result Actions
from openpyxl import load_workbook
def save_results_to_xlsx_append(results, output_dir, filename, sheet_name='K-Fold Results'):
    """
    Appends results to an existing Excel file or creates a new file if it doesn't exist.

    Args:
        results (list or pd.DataFrame): The results data to save.
        output_dir (str): The directory where the Excel file will be saved.
        filename (str): The name of the Excel file.
        sheet_name (str): The sheet name in the Excel file. Default is 'K-Fold Results'.

    Returns:
        str: The path of the saved Excel file.
    """
    # Convert results to DataFrame if necessary
    if not isinstance(results, pd.DataFrame):
        results_df = pd.DataFrame(results)
    else:
        results_df = results

    # Rearrange columns if "Identifier" is present
    if 'Identifier' in results_df.columns:
        columns_order = ['Identifier'] + [col for col in results_df.columns if col != 'Identifier']
        results_df = results_df[columns_order]

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the full output path
    output_path = os.path.join(output_dir, filename)

    # Append to existing Excel file or create a new one
    if os.path.exists(output_path):
        print(f'Appending data to existing file: {output_path}')
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the existing workbook
            existing_workbook = load_workbook(output_path)

            # Check if the sheet exists
            if sheet_name in existing_workbook.sheetnames:
                # Load existing sheet and append
                start_row = existing_workbook[sheet_name].max_row
                results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
            else:
                # Write new sheet if not exists
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        print(f'Creating new file: {output_path}')
        results_df.to_excel(output_path, index=False, sheet_name=sheet_name)

    print(f'Results successfully saved to: {output_path}')
    return output_path

# %% End Program Actions
import time
import threading
def shutdown_with_countdown(countdown_seconds=120):
    """
    Initiates a shutdown countdown, allowing the user to cancel shutdown within the given time.

    Args:
        countdown_seconds (int): The number of seconds to wait before shutting down.
    """
    def cancel_shutdown():
        nonlocal shutdown_flag
        user_input = input("\nPress 'c' and Enter to cancel shutdown: ").strip().lower()
        if user_input == 'c':
            shutdown_flag = False
            print('Shutdown cancelled.')

    # Flag to determine whether to proceed with shutdown
    shutdown_flag = True

    # Start a thread to listen for user input
    input_thread = threading.Thread(target=cancel_shutdown, daemon=True)
    input_thread.start()

    # Countdown timer
    print(f"Shutdown scheduled in {countdown_seconds} seconds. Press 'c' to cancel.")
    for i in range(countdown_seconds, 0, -1):
        print(f'Time remaining: {i} seconds', end='\r')
        time.sleep(1)

    # Check the flag after countdown
    if shutdown_flag:
        print('\nShutdown proceeding...')
        os.system('shutdown /s /t 1')  # Execute shutdown command
    else:
        print('\nShutdown aborted.')

def end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120):
    """
    Performs actions at the end of the program, such as playing a sound or shutting down the system.

    Args:
        play_sound (bool): If True, plays a notification sound.
        shutdown (bool): If True, initiates shutdown with a countdown.
        countdown_seconds (int): Countdown time for shutdown confirmation.
    """
    if play_sound:
        try:
            import winsound
            print("Playing notification sound...")
            winsound.Beep(1000, 500)  # Frequency: 1000Hz, Duration: 500ms
        except ImportError:
            print("winsound module not available. Skipping sound playback.")

    if shutdown:
        shutdown_with_countdown(countdown_seconds)

# %% Usage; Training settings
from models import models #, models_multiscale

model = models.CNN_2layers_adaptive_maxpool_3()

# %% validation 1; sfcc
fcnetwork, feature, emotion, subject_range = 'sfcc', 'PLV', 'dominance', range(1, 24)

# trainning and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, emotion_dimension=emotion, subject_range=subject_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% validation 2; cm
fcnetwork, feature, emotion, subject_range = 'cm', 'PLV', 'dominance', range(1, 24)

# trainning and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, emotion_dimension=emotion, subject_range=subject_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% validation 3; vc
fcnetwork, feature, emotion, subject_range = 'vc', 'PLV', 'dominance', range(1, 24)

# trainning and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, emotion_dimension=emotion, subject_range=subject_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% validation 4; mx
fcnetwork, feature, emotion, subject_range = 'mx', 'PLV', 'dominance', range(1, 24)

# trainning and validation
results = cnn_cross_validation_circle(model, fcnetwork, feature, emotion_dimension=emotion, subject_range=subject_range)

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'results')
filename = f"{fcnetwork}_{type(model).__name__}_{feature}.xlsx"
save_results_to_xlsx_append(results, output_dir, filename)

# %% End program actions
end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120)